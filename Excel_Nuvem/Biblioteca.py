
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
          
wb = load_workbook('Testteizinho o retorno.xlsx') #abrindo planilha
ws = wb['Sheet1'] #Selecionando a aba da planilha

def addBook(): #função para adicionar livro na planilha
    title = input('Insira o nome do livro: ').title()
    bookID = input('Insira o código do livro: ').upper()
    edition = input('Insira a edição do livro: ') + 'ª'
    author = input('Insira o nome do autor: ').title()
    volume = input('Insira o volume do livro: ')
    LCA = input('Livro, Cópia ou Anotação: ').title()
    informacoes = [disponivel,title,bookID,author,volume,edition,LCA]
    ws.append(informacoes)#Adicionar linha na planilha
    B=len(ws['A'])
    if disponivel == 'Sim':
        fill_pattern_green = PatternFill(patternType='solid',fgColor='1EFF0D')#cor da celula
        ws['A'+f'{B}'].fill = fill_pattern_green
    for col in range(1, 11):
        char = get_column_letter(col)
        currentCell = ws[char + f'{B}'] #Alinhamento da celula
        currentCell.alignment = Alignment(horizontal='center')#Alinhamento da celula
        thin_border = Border(left=Side(style='thin'), #Bordas da celula
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        ws[char + f'{B}'].border = thin_border#Bordas da celula
    print('Livro adicionado ao acervo!')

def takeBook(bookID2): #Tirar um livro da planilha
    c=0
    d= len(ws['A'])
    for row in ws.iter_rows(min_row=1, max_col=3, max_row=len(ws['A'])):
        c+=1
        for cell in row:
            if bookID2 == ws.cell(row=cell.row,column=3).value and ws.cell(row=cell.row,column=1).value=='Sim': #Se o livro estiver disponível
                print('Livro disponível')
                email = input('Insira o email: ')
                nome = input('Digite o nome: ').title()
                data = date.today()
                fill_pattern_red = PatternFill(patternType = 'solid',fgColor='FF0000')
                ws.cell(row=cell.row,column=1).value = 'Não'
                ws.cell(row=cell.row,column=1).fill = fill_pattern_red
                ws.cell(row=cell.row,column=8).value = nome
                ws.cell(row=cell.row,column=9).value = email
                ws.cell(row=cell.row,column=10).value = data
                for col in range(8,11):
                    currentCell2 = ws.cell(row=cell.row, column = col) 
                    currentCell2.alignment = Alignment(horizontal='center')
                print('Seu empréstimo foi feito')
                return
            elif bookID2 == ws.cell(row=cell.row,column=3).value and ws.cell(row=cell.row,column=1).value=='Não': # Se o livro não estiver disponivel
                print('Livro não disponível')
                return
            elif c ==  d: #Se o código do livro estiver errado ou não tiver
                print('Livro não encontrado')
                return

def cabecalho(txt):
    print('-' * 42)
    print(txt)
    print('-' * 42)
    print('1 - Adicionar Livro')
    print('2 - Retirar livro')
    print('-' * 42)

cabecalho('MENU PRINCIPAL')

escolha = int(input('Digite a opção: '))

if int(escolha) == 1:
    disponivel = 'Sim'
    addBook()
elif int(escolha) == 2:
    disponivel2 = 'Não'
    bookID2 = input('Insira o código do livro: ')
    takeBook(bookID2)
    
wb.save('Testteizinho o retorno.xlsx') #salvar alterações

